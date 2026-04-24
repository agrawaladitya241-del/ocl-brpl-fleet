"""
data_loader.py  —  OCL/BRPL Fleet Report Loader
-----------------------------------------------
Reads the contractor fleet Excel file into a tidy long-format DataFrame.

File structure:
  - 'OCL' sheet: columns = V NO, GROUP, MOB NO, then days 1–22, then LOCATION
  - 'BRPL' sheet: columns = Vehc No, GROUP, then days 1–24, then Trip count
  - Other sheets in the workbook are ignored.

The daily columns are just numbered 1, 2, 3... — we map them to days of the
user-specified month (locked to April 2026 by default).

Cells are short-form codes (e.g. ULP, TNST, LPD, MTOCL, B-M, DH, DP, ACC).
We also track whether each cell is highlighted (blue fill = route marker).
"""

from __future__ import annotations

from datetime import datetime
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


LIVE_SHEETS = ["OCL", "BRPL"]

_VEHICLE_HEADERS = {"v no", "vehc no", "vehicle no", "vehicle", "veh no"}
_GROUP_HEADERS = {"group"}
_MOBILE_HEADERS = {"mob no", "mobile no", "contact", "phone", "driver phone"}
_LOCATION_HEADERS = {"location", "loacation", "current location", "last location"}
_TRIP_COUNT_HEADERS = {"trip", "trips", "trip count", "total trip"}


def _is_highlighted(cell) -> bool:
    """Return True if cell has any non-default fill (blue route highlight)."""
    if cell is None or cell.fill is None or cell.fill.fgColor is None:
        return False
    color = cell.fill.fgColor.rgb if cell.fill.fgColor.type == "rgb" else cell.fill.fgColor.value
    if color is None:
        return False
    s = str(color)
    if s in ("00000000", "0", ""):
        return False
    return True


def _norm(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _is_day_number(value) -> Optional[int]:
    if value is None:
        return None
    try:
        n = int(str(value).strip())
        if 1 <= n <= 31:
            return n
    except (ValueError, TypeError):
        return None
    return None


def _classify_column(header_value) -> str:
    n = _norm(header_value)
    if not n:
        return "skip"
    if n in _VEHICLE_HEADERS:
        return "vehicle"
    if n in _GROUP_HEADERS:
        return "group"
    if n in _MOBILE_HEADERS:
        return "mobile"
    if n in _LOCATION_HEADERS:
        return "location"
    if n in _TRIP_COUNT_HEADERS:
        return "trip_count"
    if _is_day_number(header_value) is not None:
        return "day"
    return "skip"


def load_sheet(
    file_path_or_buffer,
    sheet_name: str,
    month: int,
    year: int,
) -> pd.DataFrame:
    """Load one contractor sheet into tidy long format."""
    wb = load_workbook(file_path_or_buffer, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return pd.DataFrame()

    ws = wb[sheet_name]
    n_cols = ws.max_column

    headers = [ws.cell(row=1, column=c).value for c in range(1, n_cols + 1)]
    col_kinds: Dict[int, str] = {}
    day_cols: List[Tuple[int, int]] = []

    for idx0, h in enumerate(headers):
        c = idx0 + 1
        kind = _classify_column(h)
        col_kinds[c] = kind
        if kind == "day":
            day_cols.append((c, _is_day_number(h)))

    def find_col(kind: str) -> Optional[int]:
        for c, k in col_kinds.items():
            if k == kind:
                return c
        return None

    vehicle_col = find_col("vehicle") or 1
    group_col = find_col("group")
    mobile_col = find_col("mobile")
    location_col = find_col("location")
    trip_col = find_col("trip_count")

    # Fallback: detect trailing column with no header but numeric values as Trip column
    # (BRPL sheet has this — col 27 is blank header but contains trip counts)
    if trip_col is None and day_cols:
        last_day_col = day_cols[-1][0]
        # Check every column AFTER the last day column for numeric-only content
        for c in range(last_day_col + 1, n_cols + 1):
            if col_kinds.get(c) not in ("skip", None):
                continue
            numeric_count = 0
            total_count = 0
            for r in range(2, min(ws.max_row + 1, 100)):  # sample first 100 rows
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    total_count += 1
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        numeric_count += 1
            if total_count >= 5 and numeric_count / total_count >= 0.8:
                trip_col = c
                col_kinds[c] = "trip_count"
                break

    if not day_cols:
        wb.close()
        return pd.DataFrame()

    records: List[Dict] = []

    for r in range(2, ws.max_row + 1):
        vehicle_raw = ws.cell(row=r, column=vehicle_col).value
        if vehicle_raw is None:
            continue
        vehicle = str(vehicle_raw).strip().upper().replace(" ", "")
        if len(vehicle) < 5 or not vehicle.startswith("OD"):
            continue

        group = ""
        if group_col is not None:
            gv = ws.cell(row=r, column=group_col).value
            if gv is not None:
                group = str(gv).strip()

        mobile = ""
        if mobile_col is not None:
            mv = ws.cell(row=r, column=mobile_col).value
            if mv is not None:
                mobile = str(mv).strip()

        location_text = ""
        if location_col is not None:
            lv = ws.cell(row=r, column=location_col).value
            if lv is not None:
                location_text = str(lv).strip()

        manual_trip: Optional[int] = None
        if trip_col is not None:
            tv = ws.cell(row=r, column=trip_col).value
            if isinstance(tv, (int, float)):
                manual_trip = int(tv)

        for c, day in day_cols:
            cell = ws.cell(row=r, column=c)
            val = cell.value
            status_raw = str(val).strip() if val is not None else ""
            highlighted = _is_highlighted(cell) if status_raw else False

            try:
                the_date = datetime(year, month, day)
            except ValueError:
                continue

            records.append({
                "vehicle": vehicle,
                "group": group,
                "mobile": mobile,
                "date": the_date,
                "status_raw": status_raw,
                "is_highlighted": highlighted,
                "location_text": location_text,
                "manual_trip_count": manual_trip,
                "contractor": sheet_name,
            })

    wb.close()
    return pd.DataFrame.from_records(records)


def load_all(
    file_path_or_buffer,
    month: int = 4,
    year: int = 2026,
    sheets: Optional[List[str]] = None,
) -> pd.DataFrame:
    """Load all live contractor sheets and concatenate."""
    if sheets is None:
        sheets = LIVE_SHEETS

    frames = []
    for name in sheets:
        try:
            df = load_sheet(file_path_or_buffer, name, month=month, year=year)
            if not df.empty:
                frames.append(df)
        except Exception:
            continue

    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)
